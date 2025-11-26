from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, ContextTypes, filters, ConversationHandler
import logging
import datetime
import os
import asyncio
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# Suppress deprecation warnings
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=UserWarning)

logging.basicConfig(level=logging.INFO)

# -------------------------------
# ADMIN CONFIGURATION
# -------------------------------
# Add your Telegram user IDs here (get your ID by messaging @userinfobot on Telegram)
ADMIN_IDS = [7722280543, 6518582583]  # Admin user IDs

# -------------------------------
# DATABASE (In-memory)
# -------------------------------
products = {}  # product_name : {price, discount, stock, photo}
sales_log = []  # list of sales dicts
profit_data = {}  # product_name : profit_margin

# Excel file path
SALES_LOG_FILE = "sales_log.xlsx"

# -------------------------------
# STATES
# -------------------------------
ADD_NAME, ADD_PRICE, ADD_DISCOUNT, ADD_STOCK, ADD_PHOTO = range(5)
ORDER_QTY, ORDER_NAME, ORDER_PHONE, ORDER_ROOM = range(4)
DELETE_ITEM = range(1)
PROFIT_ITEM, PROFIT_AMOUNT = range(2)

# -------------------------------
# HELPER FUNCTIONS
# -------------------------------
def save_sales_to_excel():
    """Save sales log to Excel file using openpyxl"""
    if not sales_log:
        return
    
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Log"
        
        # Headers
        headers = ['Date', 'Product', 'Quantity', 'Unit Price', 'Total Price', 
                   'Profit', 'Payment Method', 'Customer Name', 'Customer Phone', 
                   'Customer Room', 'User ID']
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, sale in enumerate(sales_log, 2):
            ws.cell(row=row_num, column=1, value=sale['date'])
            ws.cell(row=row_num, column=2, value=sale['product'])
            ws.cell(row=row_num, column=3, value=sale['quantity'])
            ws.cell(row=row_num, column=4, value=sale['unit_price'])
            ws.cell(row=row_num, column=5, value=sale['total_price'])
            ws.cell(row=row_num, column=6, value=sale['profit'])
            ws.cell(row=row_num, column=7, value=sale['payment_method'])
            ws.cell(row=row_num, column=8, value=sale['customer_name'])
            ws.cell(row=row_num, column=9, value=sale['customer_phone'])
            ws.cell(row=row_num, column=10, value=sale['customer_room'])
            ws.cell(row=row_num, column=11, value=sale['user_id'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(SALES_LOG_FILE)
        logging.info(f"Sales log saved to {SALES_LOG_FILE}")
    except Exception as e:
        logging.error(f"Error saving sales log: {e}")

def calculate_daily_profit():
    """Calculate total profit for today"""
    today = datetime.date.today()
    daily_profit = 0
    
    for sale in sales_log:
        sale_date = datetime.datetime.strptime(sale['date'], '%Y-%m-%d %H:%M:%S').date()
        if sale_date == today:
            daily_profit += sale.get('profit', 0)
    
    return daily_profit

def is_admin(user_id: int) -> bool:
    """Check if user is an admin"""
    return user_id in ADMIN_IDS

# -------------------------------
# START COMMAND
# -------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    # Debug logging
    logging.info(f"User ID: {user_id}, Type: {type(user_id)}")
    logging.info(f"Admin IDs: {ADMIN_IDS}, Type: {type(ADMIN_IDS[0])}")
    logging.info(f"Is Admin: {is_admin(user_id)}")
    
    # Base keyboard for all users
    keyboard = [
        [InlineKeyboardButton("🛍️ Shop", callback_data="shop")]
    ]
    
    # Add admin buttons only for authorized users
    if is_admin(user_id):
        keyboard.extend([
            [InlineKeyboardButton("➕ Add Item (Admin)", callback_data="admin_add")],
            [InlineKeyboardButton("🗑️ Delete Item (Admin)", callback_data="admin_delete")],
            [InlineKeyboardButton("💰 Add Profit (Admin)", callback_data="admin_profit")],
            [InlineKeyboardButton("📊 Daily Report (Admin)", callback_data="admin_report")]
        ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    welcome_msg = "🎉 Welcome to Hostel E-Commerce Bot!\n\n"
    
    # Debug info (remove after fixing)
    welcome_msg += f"🔍 DEBUG INFO:\n"
    welcome_msg += f"Your ID: {user_id}\n"
    welcome_msg += f"Admin IDs: {ADMIN_IDS}\n"
    welcome_msg += f"Is Admin: {is_admin(user_id)}\n\n"
    
    if is_admin(user_id):
        welcome_msg += "👑 Admin Access Granted\n\n"
    welcome_msg += "Choose an option below:"
    
    await update.message.reply_text(welcome_msg, reply_markup=reply_markup)

# -------------------------------
# ADMIN: Add Product
# -------------------------------
async def add_item_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("❌ Access denied. Admin only.")
        return ConversationHandler.END
    
    await query.message.reply_text("📝 Enter product name:")
    return ADD_NAME

async def add_item_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Access denied. Admin only.")
        return ConversationHandler.END
    
    await update.message.reply_text("📝 Enter product name:")
    return ADD_NAME

async def add_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['name'] = update.message.text
    await update.message.reply_text("💵 Enter original price:")
    return ADD_PRICE

async def add_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data['price'] = float(update.message.text)
        await update.message.reply_text("🏷️ Enter selling/discount price:")
        return ADD_DISCOUNT
    except ValueError:
        await update.message.reply_text("❌ Invalid price. Please enter a number:")
        return ADD_PRICE

async def add_discount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data['discount'] = float(update.message.text)
        await update.message.reply_text("📦 Enter stock quantity:")
        return ADD_STOCK
    except ValueError:
        await update.message.reply_text("❌ Invalid price. Please enter a number:")
        return ADD_DISCOUNT

async def add_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        stock = int(update.message.text)
        if stock <= 0:
            await update.message.reply_text("❌ Stock must be greater than 0. Items can only be added if in stock:")
            return ADD_STOCK
        context.user_data['stock'] = stock
        await update.message.reply_text("📸 Send product photo:")
        return ADD_PHOTO
    except ValueError:
        await update.message.reply_text("❌ Invalid quantity. Please enter a number:")
        return ADD_STOCK

async def add_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not update.message or not update.message.photo:
            await update.message.reply_text("❌ Please send a photo. Try again:")
            return ADD_PHOTO
        
        # Get the largest photo (last one in the list)
        photo_file = await update.message.photo[-1].get_file()
        context.user_data['photo'] = photo_file.file_id
        
        name = context.user_data['name']
        products[name] = {
            'price': context.user_data['price'],
            'discount': context.user_data['discount'],
            'stock': context.user_data['stock'],
            'photo': context.user_data['photo']
        }
        
        await update.message.reply_text(
            f"✅ Product '{name}' added successfully!\n\n"
            f"Original Price: ₹{context.user_data['price']}\n"
            f"Selling Price: ₹{context.user_data['discount']}\n"
            f"Stock: {context.user_data['stock']}"
        )
        
        # Clear user data
        context.user_data.clear()
        
        return ConversationHandler.END
    except Exception as e:
        logging.error(f"Error in add_photo: {e}")
        await update.message.reply_text(
            f"❌ Error adding photo. Please try again or use /cancel to stop.\n"
            f"Make sure you're sending an image file."
        )
        return ADD_PHOTO

# -------------------------------
# ADMIN: Delete Product
# -------------------------------
async def delete_item_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("❌ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await query.message.reply_text("❌ No products available to delete.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(f"🗑️ {name}", callback_data=f"delete_{name}")])
    keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.message.reply_text("Select product to delete:", reply_markup=reply_markup)
    return DELETE_ITEM

async def delete_item_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await update.message.reply_text("❌ No products available to delete.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(f"🗑️ {name}", callback_data=f"delete_{name}")])
    keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("Select product to delete:", reply_markup=reply_markup)
    return DELETE_ITEM

async def delete_item_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.message.reply_text("❌ Deletion cancelled.")
        return ConversationHandler.END
    
    product_name = query.data.replace("delete_", "")
    if product_name in products:
        del products[product_name]
        await query.message.reply_text(f"✅ Product '{product_name}' deleted successfully!")
    else:
        await query.message.reply_text("❌ Product not found.")
    
    return ConversationHandler.END

# -------------------------------
# ADMIN: Add Profit Margin
# -------------------------------
async def add_profit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("❌ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await query.message.reply_text("❌ No products available.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(name, callback_data=f"profit_{name}")])
    keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.message.reply_text("Select product to add profit margin:", reply_markup=reply_markup)
    return PROFIT_ITEM

async def add_profit_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await update.message.reply_text("❌ No products available.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(name, callback_data=f"profit_{name}")])
    keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("Select product to add profit margin:", reply_markup=reply_markup)
    return PROFIT_ITEM

async def profit_item_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.message.reply_text("❌ Cancelled.")
        return ConversationHandler.END
    
    product_name = query.data.replace("profit_", "")
    context.user_data['profit_item'] = product_name
    await query.message.reply_text(f"Enter profit margin for '{product_name}' (in ₹):")
    return PROFIT_AMOUNT

async def profit_amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        profit = float(update.message.text)
        product_name = context.user_data['profit_item']
        profit_data[product_name] = profit
        await update.message.reply_text(f"✅ Profit margin of ₹{profit} set for '{product_name}'")
    except ValueError:
        await update.message.reply_text("❌ Invalid amount. Please enter a number:")
        return PROFIT_AMOUNT
    
    return ConversationHandler.END

# -------------------------------
# ADMIN: Daily Report
# -------------------------------
async def daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("❌ Access denied. Admin only.")
        return
    
    daily_profit = calculate_daily_profit()
    today = datetime.date.today()
    
    today_sales = [s for s in sales_log if datetime.datetime.strptime(s['date'], '%Y-%m-%d %H:%M:%S').date() == today]
    
    report = f"📊 Daily Report - {today}\n\n"
    report += f"💰 Total Profit: ₹{daily_profit:.2f}\n"
    report += f"📦 Total Orders: {len(today_sales)}\n\n"
    
    if today_sales:
        report += "Recent Orders:\n"
        for sale in today_sales[-5:]:  # Last 5 orders
            report += f"• {sale['product']} x{sale['quantity']} - ₹{sale['total_price']:.2f}\n"
    
    await query.message.reply_text(report)
    
    # Send Excel file if exists
    if os.path.exists(SALES_LOG_FILE):
        await query.message.reply_document(document=open(SALES_LOG_FILE, 'rb'), filename=SALES_LOG_FILE)

# -------------------------------
# SHOP
# -------------------------------
async def shop_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await show_shop(query.message)

async def shop_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_shop(update.message)

async def show_shop(message):
    if not products:
        await message.reply_text("❌ No products available currently.")
        return
    
    in_stock = {name: info for name, info in products.items() if info['stock'] > 0}
    
    if not in_stock:
        await message.reply_text("❌ All products are out of stock.")
        return
    
    keyboard = []
    for name, info in in_stock.items():
        discount_text = f"₹{info['discount']}"
        if info['price'] != info['discount']:
            discount_text += f" (was ₹{info['price']})"
        keyboard.append([InlineKeyboardButton(
            f"{name} - {discount_text} | Stock: {info['stock']}", 
            callback_data=f"buy_{name}"
        )])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await message.reply_text("🛍️ Available Products:", reply_markup=reply_markup)


# -------------------------------
# ORDER PROCESS
# -------------------------------
async def buy_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    product_name = query.data.replace("buy_", "")
    
    if product_name not in products:
        await query.message.reply_text("❌ Product not found.")
        return ConversationHandler.END
    
    if products[product_name]['stock'] <= 0:
        await query.message.reply_text("❌ Product out of stock.")
        return ConversationHandler.END
    
    context.user_data['order_product'] = product_name
    
    # Show product details
    product = products[product_name]
    await query.message.reply_photo(
        photo=product['photo'],
        caption=f"🛍️ {product_name}\n\n"
                f"💵 Price: ₹{product['discount']}\n"
                f"📦 Available Stock: {product['stock']}\n\n"
                f"Enter quantity:"
    )
    
    return ORDER_QTY

async def order_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        quantity = int(update.message.text)
        product_name = context.user_data['order_product']
        product = products[product_name]
        
        if quantity <= 0:
            await update.message.reply_text("❌ Quantity must be greater than 0:")
            return ORDER_QTY
        
        if quantity > product['stock']:
            await update.message.reply_text(f"❌ Only {product['stock']} items available. Enter valid quantity:")
            return ORDER_QTY
        
        context.user_data['order_quantity'] = quantity
        total_price = product['discount'] * quantity
        context.user_data['order_total'] = total_price
        context.user_data['payment_method'] = 'Cash on Delivery'
        
        await update.message.reply_text(
            f"📋 Order Summary:\n\n"
            f"Product: {product_name}\n"
            f"Quantity: {quantity}\n"
            f"Total: ₹{total_price}\n"
            f"Payment: Cash on Delivery\n\n"
            f"📝 Enter your full name:"
        )
        
        return ORDER_NAME
    except ValueError:
        await update.message.reply_text("❌ Invalid quantity. Please enter a number:")
        return ORDER_QTY


async def order_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['customer_name'] = update.message.text
    await update.message.reply_text("📱 Enter your phone number:")
    return ORDER_PHONE

async def order_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['customer_phone'] = update.message.text
    await update.message.reply_text("🏠 Enter your hostel room number:")
    return ORDER_ROOM

async def order_room(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['customer_room'] = update.message.text
    
    # Process order
    user_id = update.effective_user.id
    product_name = context.user_data['order_product']
    quantity = context.user_data['order_quantity']
    total_price = context.user_data['order_total']
    payment_method = context.user_data['payment_method']
    
    # Update stock
    products[product_name]['stock'] -= quantity
    
    # Calculate profit
    profit = profit_data.get(product_name, 0) * quantity
    
    # Log sale
    sale_record = {
        'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'product': product_name,
        'quantity': quantity,
        'unit_price': products[product_name]['discount'],
        'total_price': total_price,
        'profit': profit,
        'payment_method': payment_method,
        'customer_name': context.user_data['customer_name'],
        'customer_phone': context.user_data['customer_phone'],
        'customer_room': context.user_data['customer_room'],
        'user_id': user_id
    }
    sales_log.append(sale_record)
    
    # Save to Excel
    save_sales_to_excel()
    
    # Send confirmation to customer
    confirmation = (
        f"✅ Order Confirmed!\n\n"
        f"📦 Product: {product_name}\n"
        f"🔢 Quantity: {quantity}\n"
        f"💰 Total: ₹{total_price}\n"
        f"💳 Payment: {payment_method}\n\n"
        f"📍 Delivery Details:\n"
        f"Name: {context.user_data['customer_name']}\n"
        f"Phone: {context.user_data['customer_phone']}\n"
        f"Room: {context.user_data['customer_room']}\n"
    )
    
    await update.message.reply_text(confirmation)
    
    # Notify all admins about the new order
    admin_notification = (
        f"🔔 NEW ORDER RECEIVED!\n\n"
        f"📦 Product: {product_name}\n"
        f"🔢 Quantity: {quantity}\n"
        f"💰 Total: ₹{total_price}\n"
        f"💵 Profit: ₹{profit}\n"
        f"💳 Payment: {payment_method}\n\n"
        f"👤 Customer Details:\n"
        f"Name: {context.user_data['customer_name']}\n"
        f"Phone: {context.user_data['customer_phone']}\n"
        f"Room: {context.user_data['customer_room']}\n"
        f"User ID: {user_id}\n\n"
        f"📊 Remaining Stock: {products[product_name]['stock']}\n"
        f"🕐 Time: {datetime.datetime.now().strftime('%I:%M %p')}"
    )
    
    # Send notification to all admins
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_notification
            )
        except Exception as e:
            logging.error(f"Failed to send notification to admin {admin_id}: {e}")
    
    return ConversationHandler.END

# -------------------------------
# CANCEL HANDLER
# -------------------------------
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Operation cancelled.")
    return ConversationHandler.END

# -------------------------------
# MAIN
# -------------------------------
async def main():
    # Replace with your bot token
    TOKEN = "7582434551:AAHoupX4K1Qv1YX_Lp2mnuty-Qxu20NzbrQ"
    
    # Build application with proper configuration for Python 3.14
    import pytz
    app = (
        ApplicationBuilder()
        .token(TOKEN)
        .concurrent_updates(True)
        .build()
    )
    
    # Add Item Conversation
    add_conv = ConversationHandler(
        entry_points=[
            CommandHandler('add_item', add_item_command),
            CallbackQueryHandler(add_item_start, pattern="^admin_add$")
        ],
        states={
            ADD_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_name)],
            ADD_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_price)],
            ADD_DISCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_discount)],
            ADD_STOCK: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_stock)],
            ADD_PHOTO: [MessageHandler(filters.PHOTO, add_photo)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Delete Item Conversation
    delete_conv = ConversationHandler(
        entry_points=[
            CommandHandler('delete_item', delete_item_command),
            CallbackQueryHandler(delete_item_start, pattern="^admin_delete$")
        ],
        states={
            DELETE_ITEM: [CallbackQueryHandler(delete_item_confirm)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Add Profit Conversation
    profit_conv = ConversationHandler(
        entry_points=[
            CommandHandler('add_profit', add_profit_command),
            CallbackQueryHandler(add_profit_start, pattern="^admin_profit$")
        ],
        states={
            PROFIT_ITEM: [CallbackQueryHandler(profit_item_selected)],
            PROFIT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, profit_amount_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Order Conversation
    order_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(buy_product, pattern="^buy_")],
        states={
            ORDER_QTY: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_quantity)],
            ORDER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_name)],
            ORDER_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_phone)],
            ORDER_ROOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_room)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("shop", shop_command))
    app.add_handler(add_conv)
    app.add_handler(delete_conv)
    app.add_handler(profit_conv)
    app.add_handler(order_conv)
    app.add_handler(CallbackQueryHandler(shop_callback, pattern="^shop$"))
    app.add_handler(CallbackQueryHandler(daily_report, pattern="^admin_report$"))
    
    print("🤖 Bot is running...")
    
    # Initialize and run
    async with app:
        await app.start()
        await app.updater.start_polling()
        
        # Keep the bot running
        try:
            await asyncio.Event().wait()
        except (KeyboardInterrupt, SystemExit):
            pass
        finally:
            await app.updater.stop()
            await app.stop()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n🛑 Bot stopped by user")