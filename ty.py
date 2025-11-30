from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, ContextTypes, filters, ConversationHandler
import logging
import datetime
import os
import asyncio
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# Suppress deprecation warnings
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=UserWarning)

logging.basicConfig(level=logging.INFO)

# -------------------------------
# ADMIN CONFIGURATION
# -------------------------------
# Add your Telegram user IDs here (get your ID by messaging @userinfobot on Telegram)
ADMIN_IDS = [7722280543, 6518582583]  # Admin user IDs

# -------------------------------
# DATABASE (In-memory)
# -------------------------------
products = {}  # product_name : {price, discount, stock, photo}
sales_log = []  # list of sales dicts
profit_data = {}  # product_name : profit_margin

# Excel file path
SALES_LOG_FILE = "sales_log.xlsx"

# Bot pause state
paused = False

# -------------------------------
# STATES
# -------------------------------
ADD_NAME, ADD_PRICE, ADD_DISCOUNT, ADD_STOCK, ADD_PHOTO = range(5)
ORDER_QTY, ORDER_NAME, ORDER_PHONE, ORDER_ROOM = range(4)
DELETE_ITEM = range(1)
PROFIT_ITEM, PROFIT_AMOUNT = range(2)
CANCEL_ORDER, CANCEL_CONFIRM = range(2)

# -------------------------------
# HELPER FUNCTIONS
# -------------------------------
def save_sales_to_excel():
    """Save sales log to Excel file using openpyxl"""
    if not sales_log:
        return
    
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Log"
        
        # Headers
        headers = ['Date', 'Product', 'Quantity', 'Unit Price', 'Total Price', 
                   'Profit', 'Payment Method', 'Customer Name', 'Customer Phone', 
                   'Customer Room', 'User ID']
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, sale in enumerate(sales_log, 2):
            ws.cell(row=row_num, column=1, value=sale['date'])
            ws.cell(row=row_num, column=2, value=sale['product'])
            ws.cell(row=row_num, column=3, value=sale['quantity'])
            ws.cell(row=row_num, column=4, value=sale['unit_price'])
            ws.cell(row=row_num, column=5, value=sale['total_price'])
            ws.cell(row=row_num, column=6, value=sale['profit'])
            ws.cell(row=row_num, column=7, value=sale['payment_method'])
            ws.cell(row=row_num, column=8, value=sale['customer_name'])
            ws.cell(row=row_num, column=9, value=sale['customer_phone'])
            ws.cell(row=row_num, column=10, value=sale['customer_room'])
            ws.cell(row=row_num, column=11, value=sale['user_id'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(SALES_LOG_FILE)
        logging.info(f"Sales log saved to {SALES_LOG_FILE}")
    except Exception as e:
        logging.error(f"Error saving sales log: {e}")

def calculate_daily_profit():
    """Calculate total profit for today"""
    today = datetime.date.today()
    daily_profit = 0
    
    for sale in sales_log:
        sale_date = datetime.datetime.strptime(sale['date'], '%Y-%m-%d %H:%M:%S').date()
        if sale_date == today:
            daily_profit += sale.get('profit', 0)
    
    return daily_profit

def is_admin(user_id: int) -> bool:
    """Check if user is an admin"""
    return user_id in ADMIN_IDS

async def check_paused(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Check if bot is paused and respond if not admin"""
    global paused
    user_id = update.effective_user.id
    if paused and not is_admin(user_id):
        if update.message:
            await update.message.reply_text("Mart is currently closed. Please try again later.")
        elif update.callback_query:
            await update.callback_query.answer()
            await update.callback_query.message.reply_text("Mart is currently closed. Please try again later.")
        return True
    return False

# -------------------------------
# ADMIN: Pause/Resume Bot
# -------------------------------
async def pause_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    global paused
    logging.info(f"pause_bot called by user {user_id}, is_admin: {is_admin(user_id)}, paused: {paused}")
    if not is_admin(user_id):
        await update.message.reply_text("âŒ Access denied. Admin only.")
        return

    if paused:
        await update.message.reply_text("âš ï¸ Bot is already paused.")
    else:
        paused = True
        await update.message.reply_text("â¸ï¸ Bot has been paused. Only admins can interact now.")

async def resume_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    global paused
    logging.info(f"resume_bot called by user {user_id}, is_admin: {is_admin(user_id)}, paused: {paused}")
    if not is_admin(user_id):
        await update.message.reply_text("âŒ Access denied. Admin only.")
        return

    if not paused:
        await update.message.reply_text("âš ï¸ Bot is not paused.")
    else:
        paused = False
        await update.message.reply_text("â–¶ï¸ Bot has been resumed. All users can interact now.")

async def check_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    is_admin_status = "Yes" if is_admin(user_id) else "No"
    await update.message.reply_text(f"Your User ID: {user_id}\nIs Admin: {is_admin_status}")

# -------------------------------
# START COMMAND
# -------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await check_paused(update, context):
        return

    user_id = update.effective_user.id
    
    # Debug logging
    logging.info(f"User ID: {user_id}, Type: {type(user_id)}")
    logging.info(f"Admin IDs: {ADMIN_IDS}, Type: {type(ADMIN_IDS[0])}")
    logging.info(f"Is Admin: {is_admin(user_id)}")
    
    # Base keyboard for all users
    keyboard = [
        [InlineKeyboardButton("ğŸ›ï¸ Shop", callback_data="shop")]
    ]
    
    # Add admin buttons only for authorized users
    if is_admin(user_id):
        keyboard.extend([
            [InlineKeyboardButton("â• Add Item (Admin)", callback_data="admin_add")],
            [InlineKeyboardButton("ğŸ—‘ï¸ Delete Item (Admin)", callback_data="admin_delete")],
            [InlineKeyboardButton("ğŸ’° Add Profit (Admin)", callback_data="admin_profit")],
            [InlineKeyboardButton("âŒ Cancel Order (Admin)", callback_data="admin_cancel")],
            [InlineKeyboardButton("ğŸ“Š Daily Report (Admin)", callback_data="admin_report")]
        ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    welcome_msg = "ğŸ‰ Welcome to Hostel E-Commerce Bot!\n\n"
    
    # Debug info (remove after fixing)
    welcome_msg += f"ğŸ” DEBUG INFO:\n"
    welcome_msg += f"Your ID: {user_id}\n"
    welcome_msg += f"Admin IDs: {ADMIN_IDS}\n"
    welcome_msg += f"Is Admin: {is_admin(user_id)}\n\n"
    
    if is_admin(user_id):
        welcome_msg += "ğŸ‘‘ Admin Access Granted\n\n"
    welcome_msg += "Choose an option below:"
    
    await update.message.reply_text(welcome_msg, reply_markup=reply_markup)

# -------------------------------
# ADMIN: Add Product
# -------------------------------
async def add_item_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END
    
    await query.message.reply_text("ğŸ“ Enter product name:")
    return ADD_NAME

async def add_item_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END
    
    await update.message.reply_text("ğŸ“ Enter product name:")
    return ADD_NAME

async def add_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['name'] = update.message.text
    await update.message.reply_text("ğŸ’µ Enter original price:")
    return ADD_PRICE

async def add_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data['price'] = float(update.message.text)
        await update.message.reply_text("ğŸ·ï¸ Enter selling/discount price:")
        return ADD_DISCOUNT
    except ValueError:
        await update.message.reply_text("âŒ Invalid price. Please enter a number:")
        return ADD_PRICE

async def add_discount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data['discount'] = float(update.message.text)
        await update.message.reply_text("ğŸ“¦ Enter stock quantity:")
        return ADD_STOCK
    except ValueError:
        await update.message.reply_text("âŒ Invalid price. Please enter a number:")
        return ADD_DISCOUNT

async def add_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        stock = int(update.message.text)
        if stock <= 0:
            await update.message.reply_text("âŒ Stock must be greater than 0. Items can only be added if in stock:")
            return ADD_STOCK
        context.user_data['stock'] = stock
        await update.message.reply_text("ğŸ“¸ Send product photo:")
        return ADD_PHOTO
    except ValueError:
        await update.message.reply_text("âŒ Invalid quantity. Please enter a number:")
        return ADD_STOCK

async def add_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not update.message or not update.message.photo:
            await update.message.reply_text("âŒ Please send a photo. Try again:")
            return ADD_PHOTO
        
        # Get the largest photo (last one in the list)
        photo_file = await update.message.photo[-1].get_file()
        context.user_data['photo'] = photo_file.file_id
        
        name = context.user_data['name']
        products[name] = {
            'price': context.user_data['price'],
            'discount': context.user_data['discount'],
            'stock': context.user_data['stock'],
            'photo': context.user_data['photo']
        }
        
        await update.message.reply_text(
            f"âœ… Product '{name}' added successfully!\n\n"
            f"Original Price: â‚¹{context.user_data['price']}\n"
            f"Selling Price: â‚¹{context.user_data['discount']}\n"
            f"Stock: {context.user_data['stock']}"
        )
        
        # Clear user data
        context.user_data.clear()
        
        return ConversationHandler.END
    except Exception as e:
        logging.error(f"Error in add_photo: {e}")
        await update.message.reply_text(
            f"âŒ Error adding photo. Please try again or use /cancel to stop.\n"
            f"Make sure you're sending an image file."
        )
        return ADD_PHOTO

# -------------------------------
# ADMIN: Delete Product
# -------------------------------
async def delete_item_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await query.message.reply_text("âŒ No products available to delete.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(f"ğŸ—‘ï¸ {name}", callback_data=f"delete_{name}")])
    keyboard.append([InlineKeyboardButton("âŒ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.message.reply_text("Select product to delete:", reply_markup=reply_markup)
    return DELETE_ITEM

async def delete_item_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await update.message.reply_text("âŒ No products available to delete.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(f"ğŸ—‘ï¸ {name}", callback_data=f"delete_{name}")])
    keyboard.append([InlineKeyboardButton("âŒ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("Select product to delete:", reply_markup=reply_markup)
    return DELETE_ITEM

async def delete_item_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.message.reply_text("âŒ Deletion cancelled.")
        return ConversationHandler.END
    
    product_name = query.data.replace("delete_", "")
    if product_name in products:
        del products[product_name]
        await query.message.reply_text(f"âœ… Product '{product_name}' deleted successfully!")
    else:
        await query.message.reply_text("âŒ Product not found.")
    
    return ConversationHandler.END

# -------------------------------
# ADMIN: Add Profit Margin
# -------------------------------
async def add_profit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await query.message.reply_text("âŒ No products available.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(name, callback_data=f"profit_{name}")])
    keyboard.append([InlineKeyboardButton("âŒ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.message.reply_text("Select product to add profit margin:", reply_markup=reply_markup)
    return PROFIT_ITEM

async def add_profit_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await update.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END
    
    if not products:
        await update.message.reply_text("âŒ No products available.")
        return ConversationHandler.END
    
    keyboard = []
    for name in products.keys():
        keyboard.append([InlineKeyboardButton(name, callback_data=f"profit_{name}")])
    keyboard.append([InlineKeyboardButton("âŒ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("Select product to add profit margin:", reply_markup=reply_markup)
    return PROFIT_ITEM

async def profit_item_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if query.data == "cancel":
        await query.message.reply_text("âŒ Cancelled.")
        return ConversationHandler.END
    
    product_name = query.data.replace("profit_", "")
    context.user_data['profit_item'] = product_name
    await query.message.reply_text(f"Enter profit margin for '{product_name}' (in â‚¹):")
    return PROFIT_AMOUNT

async def profit_amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        profit = float(update.message.text)
        product_name = context.user_data['profit_item']
        profit_data[product_name] = profit
        await update.message.reply_text(f"âœ… Profit margin of â‚¹{profit} set for '{product_name}'")
    except ValueError:
        await update.message.reply_text("âŒ Invalid amount. Please enter a number:")
        return PROFIT_AMOUNT
    
    return ConversationHandler.END

# -------------------------------
# ADMIN: Daily Report
# -------------------------------
async def daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("âŒ Access denied. Admin only.")
        return
    
    daily_profit = calculate_daily_profit()
    today = datetime.date.today()
    
    today_sales = [s for s in sales_log if datetime.datetime.strptime(s['date'], '%Y-%m-%d %H:%M:%S').date() == today]
    
    report = f"ğŸ“Š Daily Report - {today}\n\n"
    report += f"ğŸ’° Total Profit: â‚¹{daily_profit:.2f}\n"
    report += f"ğŸ“¦ Total Orders: {len(today_sales)}\n\n"
    
    if today_sales:
        report += "Recent Orders:\n"
        for sale in today_sales[-5:]:  # Last 5 orders
            report += f"â€¢ {sale['product']} x{sale['quantity']} - â‚¹{sale['total_price']:.2f}\n"
    
    await query.message.reply_text(report)
    
    # Send Excel file if exists
    if os.path.exists(SALES_LOG_FILE):
        await query.message.reply_document(document=open(SALES_LOG_FILE, 'rb'), filename=SALES_LOG_FILE)

# -------------------------------
# SHOP
# -------------------------------
async def shop_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await show_shop(query.message)

async def shop_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_shop(update.message)

async def show_shop(message):
    if not products:
        await message.reply_text("âŒ No products available currently.")
        return
    
    in_stock = {name: info for name, info in products.items() if info['stock'] > 0}
    
    if not in_stock:
        await message.reply_text("âŒ All products are out of stock.")
        return
    
    keyboard = []
    for name, info in in_stock.items():
        discount_text = f"â‚¹{info['discount']}"
        if info['price'] != info['discount']:
            discount_text += f" (was â‚¹{info['price']})"
        keyboard.append([InlineKeyboardButton(
            f"{name} - {discount_text} | Stock: {info['stock']}", 
            callback_data=f"buy_{name}"
        )])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await message.reply_text("ğŸ›ï¸ Available Products:", reply_markup=reply_markup)


# -------------------------------
# ORDER PROCESS
# -------------------------------
async def buy_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    product_name = query.data.replace("buy_", "")
    
    if product_name not in products:
        await query.message.reply_text("âŒ Product not found.")
        return ConversationHandler.END
    
    if products[product_name]['stock'] <= 0:
        await query.message.reply_text("âŒ Product out of stock.")
        return ConversationHandler.END
    
    context.user_data['order_product'] = product_name
    
    # Show product details
    product = products[product_name]
    await query.message.reply_photo(
        photo=product['photo'],
        caption=f"ğŸ›ï¸ {product_name}\n\n"
                f"ğŸ’µ Price: â‚¹{product['discount']}\n"
                f"ğŸ“¦ Available Stock: {product['stock']}\n\n"
                f"Enter quantity:"
    )
    
    return ORDER_QTY

async def order_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        quantity = int(update.message.text)
        product_name = context.user_data['order_product']
        product = products[product_name]
        
        if quantity <= 0:
            await update.message.reply_text("âŒ Quantity must be greater than 0:")
            return ORDER_QTY
        
        if quantity > product['stock']:
            await update.message.reply_text(f"âŒ Only {product['stock']} items available. Enter valid quantity:")
            return ORDER_QTY
        
        context.user_data['order_quantity'] = quantity
        total_price = product['discount'] * quantity
        context.user_data['order_total'] = total_price
        context.user_data['payment_method'] = 'Cash on Delivery'
        
        await update.message.reply_text(
            f"ğŸ“‹ Order Summary:\n\n"
            f"Product: {product_name}\n"
            f"Quantity: {quantity}\n"
            f"Total: â‚¹{total_price}\n"
            f"Payment: Cash on Delivery\n\n"
            f"ğŸ“ Enter your full name:"
        )
        
        return ORDER_NAME
    except ValueError:
        await update.message.reply_text("âŒ Invalid quantity. Please enter a number:")
        return ORDER_QTY


async def order_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['customer_name'] = update.message.text
    await update.message.reply_text("ğŸ“± Enter your phone number:")
    return ORDER_PHONE

async def order_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['customer_phone'] = update.message.text
    await update.message.reply_text("ğŸ  Enter your hostel room number:")
    return ORDER_ROOM

async def order_room(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['customer_room'] = update.message.text
    
    # Process order
    user_id = update.effective_user.id
    product_name = context.user_data['order_product']
    quantity = context.user_data['order_quantity']
    total_price = context.user_data['order_total']
    payment_method = context.user_data['payment_method']
    
    # Update stock
    products[product_name]['stock'] -= quantity
    
    # Calculate profit
    profit = profit_data.get(product_name, 0) * quantity
    
    # Log sale
    sale_record = {
        'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'product': product_name,
        'quantity': quantity,
        'unit_price': products[product_name]['discount'],
        'total_price': total_price,
        'profit': profit,
        'payment_method': payment_method,
        'customer_name': context.user_data['customer_name'],
        'customer_phone': context.user_data['customer_phone'],
        'customer_room': context.user_data['customer_room'],
        'user_id': user_id
    }
    sales_log.append(sale_record)
    
    # Save to Excel
    save_sales_to_excel()
    
    # Send confirmation to customer
    confirmation = (
        f"âœ… Order Confirmed!\n\n"
        f"ğŸ“¦ Product: {product_name}\n"
        f"ğŸ”¢ Quantity: {quantity}\n"
        f"ğŸ’° Total: â‚¹{total_price}\n"
        f"ğŸ’³ Payment: {payment_method}\n\n"
        f"ğŸ“ Delivery Details:\n"
        f"Name: {context.user_data['customer_name']}\n"
        f"Phone: {context.user_data['customer_phone']}\n"
        f"Room: {context.user_data['customer_room']}\n"
    )
    
    await update.message.reply_text(confirmation)
    
    # Notify all admins about the new order
    admin_notification = (
        f"ğŸ”” NEW ORDER RECEIVED!\n\n"
        f"ğŸ“¦ Product: {product_name}\n"
        f"ğŸ”¢ Quantity: {quantity}\n"
        f"ğŸ’° Total: â‚¹{total_price}\n"
        f"ğŸ’µ Profit: â‚¹{profit}\n"
        f"ğŸ’³ Payment: {payment_method}\n\n"
        f"ğŸ‘¤ Customer Details:\n"
        f"Name: {context.user_data['customer_name']}\n"
        f"Phone: {context.user_data['customer_phone']}\n"
        f"Room: {context.user_data['customer_room']}\n"
        f"User ID: {user_id}\n\n"
        f"ğŸ“Š Remaining Stock: {products[product_name]['stock']}\n"
        f"ğŸ• Time: {datetime.datetime.now().strftime('%I:%M %p')}"
    )
    
    # Send notification to all admins
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_notification
            )
        except Exception as e:
            logging.error(f"Failed to send notification to admin {admin_id}: {e}")
    
    return ConversationHandler.END

# -------------------------------
# ADMIN: Cancel Order
# -------------------------------
async def cancel_order_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("âŒ Access denied. Admin only.")
        return ConversationHandler.END

    if not sales_log:
        await query.message.reply_text("âŒ No orders available to cancel.")
        return ConversationHandler.END

    # Get last 10 orders
    recent_orders = sales_log[-10:]

    keyboard = []
    for i, sale in enumerate(reversed(recent_orders)):
        order_index = len(sales_log) - 1 - i
        order_text = f"{sale['date']} - {sale['product']} x{sale['quantity']} - â‚¹{sale['total_price']:.2f}"
        keyboard.append([InlineKeyboardButton(order_text, callback_data=f"cancel_order_{order_index}")])
    keyboard.append([InlineKeyboardButton("âŒ Cancel", callback_data="cancel")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.reply_text("Select order to cancel:", reply_markup=reply_markup)
    return CANCEL_ORDER

async def cancel_order_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "cancel":
        await query.message.reply_text("âŒ Cancellation cancelled.")
        return ConversationHandler.END

    order_index = int(query.data.replace("cancel_order_", ""))
    if order_index >= len(sales_log) or order_index < 0:
        await query.message.reply_text("âŒ Order not found.")
        return ConversationHandler.END

    sale = sales_log[order_index]
    context.user_data['cancel_order_index'] = order_index

    confirm_text = (
        f"âš ï¸ Confirm cancellation of this order?\n\n"
        f"ğŸ“¦ Product: {sale['product']}\n"
        f"ğŸ”¢ Quantity: {sale['quantity']}\n"
        f"ğŸ’° Total: â‚¹{sale['total_price']:.2f}\n"
        f"ğŸ‘¤ Customer: {sale['customer_name']}\n"
        f"ğŸ“± Phone: {sale['customer_phone']}\n"
        f"ğŸ  Room: {sale['customer_room']}\n\n"
        f"This will refund the stock and remove the order from logs."
    )

    keyboard = [
        [InlineKeyboardButton("âœ… Yes, Cancel Order", callback_data="confirm_cancel")],
        [InlineKeyboardButton("âŒ No, Keep Order", callback_data="cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.message.reply_text(confirm_text, reply_markup=reply_markup)
    return CANCEL_CONFIRM

async def cancel_order_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "cancel":
        await query.message.reply_text("âŒ Order cancellation cancelled.")
        return ConversationHandler.END

    order_index = context.user_data.get('cancel_order_index')
    if order_index is None or order_index >= len(sales_log) or order_index < 0:
        await query.message.reply_text("âŒ Order not found.")
        return ConversationHandler.END

    sale = sales_log.pop(order_index)  # Remove the order

    # Refund stock
    if sale['product'] in products:
        products[sale['product']]['stock'] += sale['quantity']

    # Save updated sales log
    save_sales_to_excel()

    # Notify customer
    try:
        await context.bot.send_message(
            chat_id=sale['user_id'],
            text=(
                f"âŒ Your order has been cancelled by admin.\n\n"
                f"ğŸ“¦ Product: {sale['product']}\n"
                f"ğŸ”¢ Quantity: {sale['quantity']}\n"
                f"ğŸ’° Refunded Amount: â‚¹{sale['total_price']:.2f}\n\n"
                f"If you have any questions, please contact support."
            )
        )
    except Exception as e:
        logging.error(f"Failed to notify customer {sale['user_id']}: {e}")

    await query.message.reply_text(f"âœ… Order cancelled successfully. Stock refunded for '{sale['product']}'.")
    return ConversationHandler.END

# -------------------------------
# CANCEL HANDLER
# -------------------------------
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("âŒ Operation cancelled.")
    return ConversationHandler.END

# -------------------------------
# MAIN
# -------------------------------
async def main():
    # Replace with your bot token
    TOKEN = "7582434551:AAFhNAkMO7OiOsG3gWoZpGKLhyaV0uvdzEw"
    
    # Build application with proper configuration for Python 3.14
    import pytz
    app = (
        ApplicationBuilder()
        .token(TOKEN)
        .concurrent_updates(True)
        .build()
    )
    
    # Add Item Conversation
    add_conv = ConversationHandler(
        entry_points=[
            CommandHandler('add_item', add_item_command),
            CallbackQueryHandler(add_item_start, pattern="^admin_add$")
        ],
        states={
            ADD_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_name)],
            ADD_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_price)],
            ADD_DISCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_discount)],
            ADD_STOCK: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_stock)],
            ADD_PHOTO: [MessageHandler(filters.PHOTO, add_photo)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Delete Item Conversation
    delete_conv = ConversationHandler(
        entry_points=[
            CommandHandler('delete_item', delete_item_command),
            CallbackQueryHandler(delete_item_start, pattern="^admin_delete$")
        ],
        states={
            DELETE_ITEM: [CallbackQueryHandler(delete_item_confirm)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Add Profit Conversation
    profit_conv = ConversationHandler(
        entry_points=[
            CommandHandler('add_profit', add_profit_command),
            CallbackQueryHandler(add_profit_start, pattern="^admin_profit$")
        ],
        states={
            PROFIT_ITEM: [CallbackQueryHandler(profit_item_selected)],
            PROFIT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, profit_amount_entered)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Cancel Order Conversation
    cancel_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cancel_order_start, pattern="^admin_cancel$")],
        states={
            CANCEL_ORDER: [CallbackQueryHandler(cancel_order_select)],
            CANCEL_CONFIRM: [CallbackQueryHandler(cancel_order_confirm)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Order Conversation
    order_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(buy_product, pattern="^buy_")],
        states={
            ORDER_QTY: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_quantity)],
            ORDER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_name)],
            ORDER_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_phone)],
            ORDER_ROOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, order_room)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("shop", shop_command))
    app.add_handler(CommandHandler("pause", pause_bot))
    app.add_handler(CommandHandler("resume", resume_bot))
    app.add_handler(CommandHandler("check_id", check_id))
    app.add_handler(add_conv)
    app.add_handler(delete_conv)
    app.add_handler(profit_conv)
    app.add_handler(cancel_conv)
    app.add_handler(order_conv)
    app.add_handler(CallbackQueryHandler(shop_callback, pattern="^shop$"))
    app.add_handler(CallbackQueryHandler(daily_report, pattern="^admin_report$"))
    
    print("ğŸ¤– Bot is running...")
    
    # Initialize and run
    async with app:
        await app.start()
        await app.updater.start_polling()
        
        # Keep the bot running
        try:
            await asyncio.Event().wait()
        except (KeyboardInterrupt, SystemExit):
            pass
        finally:
            await app.updater.stop()
            await app.stop()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nğŸ›‘ Bot stopped by user")